# 给定excel表格，看NAS中pass/preprocessing/segmentation 有没有对应ID


# -*- coding: utf-8 -*-
"""
检查 NeedSeg_20260604.xlsx 里每个 site 的 ID，
是否存在于 NAS 对应 site 的 3_sMRI_preprocessing 文件夹中。

运行环境：Windows Python
依赖：openpyxl
安装：pip install openpyxl
"""

from __future__ import annotations

import os
import subprocess
from pathlib import Path
from collections import OrderedDict

from openpyxl import load_workbook


# =========================
# 你最常改的配置区
# =========================

EXCEL_PATH = r"D:\University\master\QC2026\2026\NeedSeg_20260604.xlsx"

NAS_IP = "10.19.136.231"
NAS_SHARE = rf"\\{NAS_IP}\002"

NAS_USER = "linm"
NAS_PASSWORDS = [
    "fGWH|1,I",
    "l9cG5/g{",
]

BASE_DIR = rf"{NAS_SHARE}\CBCP\CBCP_MRI\4_VisualQC_Processing"

SITE_CONFIGS = [
    {
        "site_key": "site1",
        "site_name": "STU",
        "folder": "Site1_STU",
        "sheet_keywords": ["site1", "stu"],
    },
    {
        "site_key": "site2",
        "site_name": "XMH",
        "folder": "Site2_XMH",
        "sheet_keywords": ["site2", "xmh"],
    },
    {
        "site_key": "site3",
        "site_name": "CZH",
        "folder": "Site3_CZH",
        "sheet_keywords": ["site3", "czh"],
    },
]

# 如果 Excel 里 010xxx 被读成 10xxx，可以用“去前导 0”兜底匹配。
# 建议保持 True。
MATCH_STRIP_LEADING_ZEROS = True

# 是否只看文件夹。一般 preprocessing 下面都是被试文件夹，所以 True。
ONLY_DIRECTORIES = True


# =========================
# 工具函数
# =========================

def run_cmd(args: list[str]) -> tuple[int, str]:
    """
    运行 Windows 命令，不使用 shell=True，避免密码里的 | { / 等特殊字符炸掉。
    """
    p = subprocess.run(
        args,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="gbk",
        errors="replace",
        shell=False,
    )
    return p.returncode, p.stdout


def disconnect_nas() -> None:
    """
    清理已有连接，避免 Windows 1219 错误：
    一个用户不能用不同账号/密码重复连接同一服务器。
    """
    run_cmd(["net", "use", NAS_SHARE, "/delete", "/y"])


def connect_nas() -> str:
    """
    依次尝试两个密码，成功后返回成功使用的密码。
    """
    last_output = ""

    for pwd in NAS_PASSWORDS:
        disconnect_nas()

        code, out = run_cmd([
            "net", "use",
            NAS_SHARE,
            pwd,
            f"/user:{NAS_USER}",
            "/persistent:no",
        ])

        last_output = out

        if code == 0:
            print(f"[OK] NAS 连接成功：{NAS_SHARE}，使用密码：{pwd}")
            return pwd

        print(f"[WARN] 密码尝试失败：{pwd}")
        print(out.strip())

    raise RuntimeError(
        "两个密码都连接失败。最后一次 net use 输出如下：\n"
        + last_output
    )


def clean_cell_value(value) -> str | None:
    """
    把 Excel 单元格值转成干净字符串。
    """
    if value is None:
        return None

    if isinstance(value, float):
        if value.is_integer():
            s = str(int(value))
        else:
            s = str(value)
    else:
        s = str(value)

    s = s.strip()

    if not s:
        return None

    lower = s.lower()
    if lower in {"nan", "none", "null"}:
        return None

    # 有些 Excel 会读成 "123.0"
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]

    # 去掉首尾引号
    s = s.strip("'\"")

    return s if s else None


def extract_id(raw_name: str, site_key: str) -> str | None:
    """
    从 Excel ID 或 NAS 文件夹名中提取用于比较的 ID。

    site1/site3:
        ID_xxmo -> ID

    site2:
        Nxxx_xxx_xxmo -> Nxxx
        用户已说明 Nxxx unique，所以只取第一个下划线前的部分。
    """
    s = clean_cell_value(raw_name)
    if not s:
        return None

    # 如果路径不小心传进来，只取最后一级
    s = os.path.basename(s)

    # 所有 site 都按第一个下划线前作为主 ID。
    # 对 site2 来说，这正好得到 Nxxx。
    if "_" in s:
        s = s.split("_", 1)[0]

    s = s.strip()
    return s if s else None


def canonical_id(s: str) -> str:
    """
    比较用 canonical key。
    Windows 文件名大小写不敏感，所以这里统一大写。
    """
    return s.strip().upper()


def resolve_sheet_name(wb, site_cfg: dict, fallback_index: int) -> str:
    """
    优先根据 sheet 名里的关键词匹配 site；
    匹配不到就用第几个 sheet 兜底。
    """
    sheet_names = wb.sheetnames
    keywords = [k.lower() for k in site_cfg["sheet_keywords"]]

    # 1. 精确匹配
    for name in sheet_names:
        if name.lower() in keywords:
            return name

    # 2. 包含匹配
    for name in sheet_names:
        lname = name.lower()
        if any(k in lname for k in keywords):
            return name

    # 3. 兜底：按顺序 site1/site2/site3 对应前三个 sheet
    if fallback_index < len(sheet_names):
        return sheet_names[fallback_index]

    raise RuntimeError(
        f"找不到 {site_cfg['site_key']} 对应的 sheet，"
        f"当前工作簿 sheets = {sheet_names}"
    )


def read_excel_ids(wb, sheet_name: str, site_key: str) -> list[str]:
    """
    读取某个 sheet 第一列，从第二行开始。
    保留原始顺序，自动去重。
    """
    ws = wb[sheet_name]

    ids = []
    seen = set()

    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        sid = extract_id(value, site_key)

        if sid is None:
            continue

        key = canonical_id(sid)
        if key not in seen:
            seen.add(key)
            ids.append(sid)

    return ids


def scan_nas_site_ids(site_cfg: dict) -> tuple[set[str], dict[str, str]]:
    """
    扫描 NAS 对应 site 的 preprocessing 目录。
    返回：
        folder_keys: 用于直接匹配的 canonical ID set
        nozero_map: 去前导 0 后的 ID -> 原 ID，用于兜底匹配
    """
    site_key = site_cfg["site_key"]
    folder = site_cfg["folder"]

    site_dir = os.path.join(
        BASE_DIR,
        folder,
        "1_sMRI",
        "3_sMRI_preprocessing",
    )

    if not os.path.exists(site_dir):
        raise FileNotFoundError(f"NAS 路径不存在：{site_dir}")

    folder_keys = set()
    nozero_map = {}

    total_items = 0
    used_items = 0

    with os.scandir(site_dir) as it:
        for entry in it:
            total_items += 1

            if ONLY_DIRECTORIES and not entry.is_dir():
                continue

            sid = extract_id(entry.name, site_key)
            if sid is None:
                continue

            used_items += 1

            key = canonical_id(sid)
            folder_keys.add(key)

            nozero_key = canonical_id(sid.lstrip("0") or "0")
            if nozero_key not in nozero_map:
                nozero_map[nozero_key] = sid

    print(f"[INFO] {site_cfg['site_key']} {site_cfg['site_name']} NAS 扫描完成：")
    print(f"       路径：{site_dir}")
    print(f"       总项目数：{total_items}")
    print(f"       纳入比较数：{used_items}")
    print(f"       唯一 ID 数：{len(folder_keys)}")

    return folder_keys, nozero_map


def find_missing_ids(excel_ids: list[str], folder_keys: set[str], nozero_map: dict[str, str]) -> tuple[list[str], list[tuple[str, str]]]:
    """
    返回：
        missing: 真正没匹配到的 Excel ID
        matched_by_nozero: 通过去前导 0 兜底匹配到的 ID
    """
    missing = []
    matched_by_nozero = []

    for sid in excel_ids:
        key = canonical_id(sid)

        if key in folder_keys:
            continue

        if MATCH_STRIP_LEADING_ZEROS:
            nozero_key = canonical_id(sid.lstrip("0") or "0")
            if nozero_key in nozero_map:
                matched_by_nozero.append((sid, nozero_map[nozero_key]))
                continue

        missing.append(sid)

    return missing, matched_by_nozero


def main() -> None:
    excel_path = Path(EXCEL_PATH)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在：{excel_path}")

    print("=" * 80)
    print("开始检查 NeedSeg Excel ID 是否存在于 NAS preprocessing 文件夹")
    print("=" * 80)

    connect_nas()

    wb = load_workbook(excel_path, read_only=True, data_only=True)

    print("\n[INFO] Excel sheets：")
    for i, name in enumerate(wb.sheetnames, start=1):
        print(f"       {i}. {name}")

    all_results = OrderedDict()

    for idx, site_cfg in enumerate(SITE_CONFIGS):
        site_key = site_cfg["site_key"]
        site_name = site_cfg["site_name"]

        print("\n" + "-" * 80)
        print(f"检查 {site_key} {site_name}")
        print("-" * 80)

        sheet_name = resolve_sheet_name(wb, site_cfg, fallback_index=idx)
        print(f"[INFO] 使用 sheet：{sheet_name}")

        excel_ids = read_excel_ids(wb, sheet_name, site_key)
        print(f"[INFO] Excel 读取到唯一 ID 数：{len(excel_ids)}")

        folder_keys, nozero_map = scan_nas_site_ids(site_cfg)

        missing, matched_by_nozero = find_missing_ids(
            excel_ids=excel_ids,
            folder_keys=folder_keys,
            nozero_map=nozero_map,
        )

        all_results[site_key] = {
            "site_name": site_name,
            "sheet_name": sheet_name,
            "excel_count": len(excel_ids),
            "nas_count": len(folder_keys),
            "missing": missing,
            "matched_by_nozero": matched_by_nozero,
        }

        if matched_by_nozero:
            print(f"[WARN] 有 {len(matched_by_nozero)} 个 ID 通过去前导 0 兜底匹配到了：")
            for excel_id, nas_id in matched_by_nozero[:20]:
                print(f"       Excel: {excel_id}  ->  NAS: {nas_id}")
            if len(matched_by_nozero) > 20:
                print(f"       ……还有 {len(matched_by_nozero) - 20} 个未显示")

        if missing:
            print(f"[FAIL] {site_key} {site_name} 有 {len(missing)} 个 Excel ID 在 NAS 没搜到：")
            for sid in missing:
                print(f"       {sid}")
        else:
            print(f"[OK] {site_key} {site_name}：Excel 里的 ID 在 NAS 都能找到。")

    print("\n" + "=" * 80)
    print("总汇总")
    print("=" * 80)

    total_missing = 0

    for site_key, result in all_results.items():
        missing = result["missing"]
        total_missing += len(missing)

        print(
            f"{site_key} {result['site_name']} | "
            f"sheet={result['sheet_name']} | "
            f"Excel={result['excel_count']} | "
            f"NAS={result['nas_count']} | "
            f"Missing={len(missing)}"
        )

    print("-" * 80)

    if total_missing == 0:
        print("[OK] 所有 site 都没有缺失 ID。")
    else:
        print(f"[FAIL] 总共有 {total_missing} 个 ID 在对应 NAS site 中没搜到。")
        print("上面每个 site 的 [FAIL] 部分已经逐个打印。")


if __name__ == "__main__":
    main()