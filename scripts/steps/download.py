# -*- coding: utf-8 -*-
"""
CBCP MRI QC 数据检查与复制脚本

输出逻辑：
1. 终端只打印每个 site 的 summary，不再刷屏。
2. 详细日志写入 PROJECT_DIR/qc_scan_log.txt，每次运行覆盖。
3. 日志按 SITE 分组：
   - 先查 3_sMRI_preprocessing：列出哪些 ID 找到、哪些没找到、哪些多候选。
   - 只有 3_sMRI_preprocessing 没找到的，才继续查 1_sMRI_data。
   - 再列出 raw 找到、raw 没找到、raw 多候选。
4. 默认 DRY RUN，只判断和写日志，不复制。
"""

import os
import re
import csv
import subprocess
from pathlib import Path
from dataclasses import dataclass
from collections import defaultdict
from typing import Optional, Dict, List, Tuple, Any

from openpyxl import load_workbook


# =========================
# 你通常只需要改这里
# =========================

EXECUTE_COPY = True

EXCEL_PATH = r"D:\University\master\QC2026\2026\table.xlsx"

PROJECT_DIR = Path(r"D:\University\master\QC2026\2026\0604_CBCP")

LOCAL_RAW_DIR = PROJECT_DIR / "0_rawdata"
LOCAL_RESULTS_DIR = PROJECT_DIR / "6_results"

LOG_PATH = PROJECT_DIR / "qc_scan_log.txt"
REPORT_CSV_PATH = PROJECT_DIR / "qc_scan_report.csv"

NAS_SHARE = r"\\10.19.136.231\002"
NAS_USERNAME = "linm"
NAS_PASSWORD = "l9cG5/g{"

REMOTE_BASE = Path(
    r"\\10.19.136.231\002\CBCP\CBCP_MRI\4_VisualQC_Processing"
)

SHEET_SITE_MAP = {
    "site1_stu": "Site1_STU",
    "site2_xmh": "Site2_XMH",
    "site3_czh": "Site3_CZH",
}

# 你的目录结构就是目标目录下一层文件夹，所以这里用 1。
MAX_SEARCH_DEPTH = 1

# 默认关闭 contains，避免误匹配。
# 现在核心逻辑是：文件夹名去掉末尾 _数字mo 后，与表格 ID 比较。
ALLOW_CONTAINS_FALLBACK = False

SKIP_IF_DEST_EXISTS = True

MAX_CANDIDATES_IN_LOG = 30


# =========================
# 日志工具
# =========================

_LOG_FH = None


def init_log() -> None:
    global _LOG_FH
    PROJECT_DIR.mkdir(parents=True, exist_ok=True)
    _LOG_FH = open(LOG_PATH, "w", encoding="utf-8-sig", newline="\n")


def close_log() -> None:
    global _LOG_FH
    if _LOG_FH is not None:
        _LOG_FH.close()
        _LOG_FH = None


def log(msg: str = "") -> None:
    if _LOG_FH is not None:
        _LOG_FH.write(str(msg) + "\n")
        _LOG_FH.flush()


def log_section(title: str) -> None:
    log("")
    log("=" * 120)
    log(title)
    log("=" * 120)


def log_subsection(title: str) -> None:
    log("")
    log("-" * 120)
    log(title)
    log("-" * 120)


# =========================
# ID 处理
# =========================

def strip_age_suffix(name: str) -> str:
    """
    只剥离末尾月龄。

    0101010001_75mo      -> 0101010001
    0101010002_6mo       -> 0101010002
    N1013_xxxxxx_3mo     -> N1013_xxxxxx

    注意：SITE2/XMH 可能有两个下划线，所以不能 split("_")[0]。
    """
    s = str(name).strip()
    s = re.sub(r"[_\-\s]*\d+(?:\.\d+)?\s*mo$", "", s, flags=re.IGNORECASE)
    return s.strip()


def normalize_key(s: str) -> str:
    s = strip_age_suffix(str(s))
    s = s.strip().lower()
    s = re.sub(r"[^0-9a-zA-Z]+", "", s)
    return s


def digits_only(s: str) -> str:
    s = strip_age_suffix(str(s))
    return re.sub(r"\D+", "", s)


def make_id_variants(sub_id: str) -> List[str]:
    """
    给 Excel ID 做少量变体，主要防止 Excel 吃掉前导 0。
    正常情况下，优先用原始 ID 精确匹配。
    """
    raw = str(sub_id).strip()

    if re.fullmatch(r"\d+\.0", raw):
        raw = raw[:-2]

    core = strip_age_suffix(raw)

    variants = []

    def add(x: Any) -> None:
        x = str(x).strip()
        if x and x not in variants:
            variants.append(x)

    add(raw)
    add(core)
    add(core.lower())
    add(normalize_key(core))

    d = digits_only(core)

    if d:
        add(d)

        try:
            n = int(d)
            for width in [4, 5, 6, 7, 8, 9, 10, 11, 12]:
                if len(d) <= width:
                    add(f"{n:0{width}d}")
        except ValueError:
            pass

    current = list(variants)
    for v in current:
        add(strip_age_suffix(v))
        add(normalize_key(v))

    return variants


def cell_to_id_and_debug(cell) -> Tuple[Optional[str], Dict[str, Any]]:
    v = cell.value

    debug = {
        "raw_value": repr(v),
        "python_type": type(v).__name__,
        "number_format": str(cell.number_format),
        "data_type": str(cell.data_type),
    }

    if v is None:
        return None, debug

    if isinstance(v, str):
        sub_id = v.strip()
        if re.fullmatch(r"\d+\.0", sub_id):
            sub_id = sub_id[:-2]
        sub_id = strip_age_suffix(sub_id)
        return (sub_id if sub_id else None), debug

    if isinstance(v, int):
        fmt = str(cell.number_format)
        if re.fullmatch(r"0+", fmt):
            return f"{v:0{len(fmt)}d}", debug
        return str(v).strip(), debug

    if isinstance(v, float):
        if v.is_integer():
            fmt = str(cell.number_format)
            if re.fullmatch(r"0+", fmt):
                return f"{int(v):0{len(fmt)}d}", debug
            return str(int(v)).strip(), debug
        return str(v).strip(), debug

    return strip_age_suffix(str(v).strip()), debug


def resolve_site_folder(sheet_name: str) -> Optional[str]:
    key = sheet_name.strip().lower()

    if key in SHEET_SITE_MAP:
        return SHEET_SITE_MAP[key]

    if "stu" in key:
        return "Site1_STU"
    if "xmh" in key:
        return "Site2_XMH"
    if "czh" in key:
        return "Site3_CZH"

    return None


# =========================
# 文件夹索引
# =========================

@dataclass
class FolderEntry:
    name: str
    path: Path
    rel_path: str
    depth: int
    core_id: str
    core_norm: str
    core_digits: str


@dataclass
class FolderIndex:
    parent: Path
    entries: List[FolderEntry]
    core_lower_map: Dict[str, List[FolderEntry]]
    core_norm_map: Dict[str, List[FolderEntry]]
    core_digits_map: Dict[str, List[FolderEntry]]


def build_folder_index(parent: Path, max_depth: int) -> FolderIndex:
    """
    max_depth=1：
        只遍历 parent 下面一层文件夹。
        例如：
        parent/0101010001_75mo
        parent/N1013_0301010538_3mo
    """
    if not parent.exists():
        raise FileNotFoundError(f"NAS 目录不存在：{parent}")

    entries: List[FolderEntry] = []
    core_lower_map: Dict[str, List[FolderEntry]] = {}
    core_norm_map: Dict[str, List[FolderEntry]] = {}
    core_digits_map: Dict[str, List[FolderEntry]] = {}

    parent = Path(parent)

    for root, dirs, files in os.walk(parent):
        root_path = Path(root)

        try:
            rel_root = root_path.relative_to(parent)
            root_depth = 0 if str(rel_root) == "." else len(rel_root.parts)
        except ValueError:
            root_depth = 0

        if root_depth >= max_depth:
            dirs[:] = []
            continue

        for d in dirs:
            p = root_path / d

            try:
                rel_path = str(p.relative_to(parent))
                depth = len(Path(rel_path).parts)
            except ValueError:
                rel_path = d
                depth = 1

            if depth > max_depth:
                continue

            core_id = strip_age_suffix(d)
            core_norm = normalize_key(core_id)
            core_digits = digits_only(core_id)

            e = FolderEntry(
                name=d,
                path=p,
                rel_path=rel_path,
                depth=depth,
                core_id=core_id,
                core_norm=core_norm,
                core_digits=core_digits,
            )

            entries.append(e)
            core_lower_map.setdefault(core_id.lower(), []).append(e)

            if core_norm:
                core_norm_map.setdefault(core_norm, []).append(e)

            if core_digits:
                core_digits_map.setdefault(core_digits, []).append(e)

    return FolderIndex(
        parent=parent,
        entries=entries,
        core_lower_map=core_lower_map,
        core_norm_map=core_norm_map,
        core_digits_map=core_digits_map,
    )


def unique_entries(entries: List[FolderEntry]) -> List[FolderEntry]:
    seen = set()
    out = []

    for e in entries:
        key = str(e.path).lower()
        if key not in seen:
            seen.add(key)
            out.append(e)

    return out


def find_id_folder(index: FolderIndex, sub_id: str) -> Tuple[Optional[FolderEntry], str, List[FolderEntry]]:
    sub_id = str(sub_id).strip()

    if not sub_id:
        return None, "empty_id", []

    variants = make_id_variants(sub_id)

    # 1. core_id 精确匹配
    candidates: List[FolderEntry] = []

    for v in variants:
        core = strip_age_suffix(v)
        hits = index.core_lower_map.get(core.lower(), [])
        candidates.extend(hits)

    candidates = unique_entries(candidates)

    if len(candidates) == 1:
        return candidates[0], "core_exact", candidates

    if len(candidates) > 1:
        return None, "multiple_core_exact", candidates

    # 2. normalize 后匹配
    candidates = []

    for v in variants:
        nv = normalize_key(v)
        if not nv:
            continue
        hits = index.core_norm_map.get(nv, [])
        candidates.extend(hits)

    candidates = unique_entries(candidates)

    if len(candidates) == 1:
        return candidates[0], "core_normalized", candidates

    if len(candidates) > 1:
        return None, "multiple_core_normalized", candidates

    # 3. 纯数字匹配，主要防止 Excel 吃前导 0
    candidates = []

    for v in variants:
        dv = digits_only(v)
        if not dv:
            continue
        hits = index.core_digits_map.get(dv, [])
        candidates.extend(hits)

    candidates = unique_entries(candidates)

    if len(candidates) == 1:
        return candidates[0], "core_digits", candidates

    if len(candidates) > 1:
        return None, "multiple_core_digits", candidates

    # 4. 可选 contains 兜底，默认关闭
    if ALLOW_CONTAINS_FALLBACK:
        candidates = []

        variant_norms = []
        for v in variants:
            nv = normalize_key(v)
            if len(nv) >= 4:
                variant_norms.append(nv)

        variant_norms = list(dict.fromkeys(variant_norms))

        for e in index.entries:
            for nv in variant_norms:
                if nv in e.core_norm or e.core_norm in nv:
                    candidates.append(e)

        candidates = unique_entries(candidates)

        if len(candidates) == 1:
            return candidates[0], "core_contains_fallback", candidates

        if len(candidates) > 1:
            return None, "multiple_core_contains_fallback", candidates

    return None, "not_found", []


# =========================
# NAS 和复制
# =========================

def run_net_use() -> None:
    commands = [
        ["net", "use", NAS_SHARE, f"/user:{NAS_USERNAME}", NAS_PASSWORD, "/persistent:no"],
        ["net", "use", NAS_SHARE, f"/user:10.19.136.231\\{NAS_USERNAME}", NAS_PASSWORD, "/persistent:no"],
    ]

    log_section("[NAS] 连接 NAS")
    log(f"NAS_SHARE    : {NAS_SHARE}")
    log(f"NAS_USERNAME : {NAS_USERNAME}")
    log("NAS_PASSWORD : ******")

    last_stdout = ""
    last_stderr = ""

    for cmd in commands:
        safe_cmd = cmd.copy()
        safe_cmd[4] = "******"
        log(f"[NAS] 尝试命令：{' '.join(safe_cmd)}")

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding="gbk",
            errors="ignore",
            shell=False
        )

        last_stdout = result.stdout.strip()
        last_stderr = result.stderr.strip()

        log(f"[NAS] returncode: {result.returncode}")

        if last_stdout:
            log(f"[NAS] stdout: {last_stdout}")

        if last_stderr:
            log(f"[NAS] stderr: {last_stderr}")

        if result.returncode == 0:
            log("[NAS] 连接成功。")
            return

        if Path(NAS_SHARE).exists():
            log("[NAS] 当前 NAS 路径已经可访问，继续执行。")
            return

    log("[NAS] 连接失败。")
    log(r"可以先在 CMD 执行：net use \\10.19.136.231\002 /delete")
    log(f"最后 stdout: {last_stdout}")
    log(f"最后 stderr: {last_stderr}")

    raise RuntimeError("无法连接 NAS。")


def robocopy_folder(src: Path, dst: Path) -> Tuple[bool, str]:
    if SKIP_IF_DEST_EXISTS and dst.exists():
        return True, "SKIPPED_DEST_EXISTS"

    dst.parent.mkdir(parents=True, exist_ok=True)

    cmd = [
        "robocopy",
        str(src),
        str(dst),
        "/E",
        "/COPY:DAT",
        "/DCOPY:DAT",
        "/R:2",
        "/W:2",
        "/MT:16",
        "/NP",
    ]

    log(f"[COPY] command: {' '.join(cmd)}")

    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="gbk",
        errors="ignore",
        shell=False
    )

    log(f"[COPY] returncode: {result.returncode}")

    if result.stdout:
        log("[COPY] stdout:")
        log(result.stdout)

    if result.stderr:
        log("[COPY] stderr:")
        log(result.stderr)

    if result.returncode <= 7:
        return True, f"ROBOCOPY_OK_CODE_{result.returncode}"

    return False, f"ROBOCOPY_FAILED_CODE_{result.returncode}"


# =========================
# Excel
# =========================

def load_ids_from_excel(excel_path: Path) -> List[Dict[str, Any]]:
    log_section("[EXCEL] 读取表格")
    log(f"Excel path: {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)

    records = []

    for ws in wb.worksheets:
        sheet_name = ws.title
        site_folder = resolve_site_folder(sheet_name)

        if site_folder is None:
            log(f"[SKIP] sheet 无法识别来源，跳过：{sheet_name}")
            continue

        header = ws.cell(row=1, column=1).value
        log(f"[EXCEL] sheet={sheet_name}, site={site_folder}, A1={header}")

        sheet_count = 0

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            sub_id, cell_debug = cell_to_id_and_debug(cell)

            if sub_id is None:
                continue

            records.append({
                "sheet": sheet_name,
                "site_folder": site_folder,
                "row": row_idx,
                "sub_id": sub_id,
                "cell_debug": cell_debug,
            })

            sheet_count += 1

        log(f"[EXCEL] sheet={sheet_name} 读到 ID 数：{sheet_count}")

    return records


# =========================
# 分组日志输出
# =========================

def format_candidates(candidates: List[FolderEntry]) -> str:
    parts = []
    for e in candidates[:MAX_CANDIDATES_IN_LOG]:
        parts.append(f"{e.rel_path} [core_id={e.core_id}]")

    if len(candidates) > MAX_CANDIDATES_IN_LOG:
        parts.append(f"... 其余 {len(candidates) - MAX_CANDIDATES_IN_LOG} 个候选省略")

    return "; ".join(parts)


def log_site_index_info(site_folder: str, processed_index: FolderIndex, raw_index: FolderIndex) -> None:
    log_subsection(f"[{site_folder}] NAS 索引信息")
    log(f"3_sMRI_preprocessing path : {processed_index.parent}")
    log(f"3_sMRI_preprocessing folders: {len(processed_index.entries)}")
    log(f"1_sMRI_data path          : {raw_index.parent}")
    log(f"1_sMRI_data folders       : {len(raw_index.entries)}")

    log("")
    log("3_sMRI_preprocessing 前 20 个文件夹样例：")
    for e in processed_index.entries[:20]:
        log(f"  folder={e.rel_path} | core_id={e.core_id}")

    log("")
    log("1_sMRI_data 前 20 个文件夹样例：")
    for e in raw_index.entries[:20]:
        log(f"  folder={e.rel_path} | core_id={e.core_id}")


def log_site_step1(site_folder: str, total_n: int, found: List[Dict[str, Any]],
                   not_found: List[Dict[str, Any]], ambiguous: List[Dict[str, Any]]) -> None:
    log_subsection(f"[{site_folder}] STEP 1：查 3_sMRI_preprocessing")

    log(f"Excel ID 总数: {total_n}")
    log(f"FOUND_IN_3_PREPROCESSING: {len(found)}")
    for item in found:
        rec = item["rec"]
        entry = item["entry"]
        log(
            f"  [FOUND_3] row={rec['row']} | ID={rec['sub_id']} "
            f"-> folder={entry.rel_path} | core_id={entry.core_id} | match={item['match']}"
        )

    log("")
    log(f"NOT_FOUND_IN_3_PREPROCESSING_GO_RAW: {len(not_found)}")
    for item in not_found:
        rec = item["rec"]
        log(f"  [NOT_FOUND_3] row={rec['row']} | ID={rec['sub_id']} | match={item['match']}")

    log("")
    log(f"AMBIGUOUS_IN_3_PREPROCESSING: {len(ambiguous)}")
    for item in ambiguous:
        rec = item["rec"]
        log(
            f"  [AMBIGUOUS_3] row={rec['row']} | ID={rec['sub_id']} "
            f"| match={item['match']} | candidates={format_candidates(item['candidates'])}"
        )


def log_site_step2(site_folder: str, checked_n: int, found: List[Dict[str, Any]],
                   missing: List[Dict[str, Any]], ambiguous: List[Dict[str, Any]]) -> None:
    log_subsection(f"[{site_folder}] STEP 2：对 STEP 1 没找到的 ID 查 1_sMRI_data")

    log(f"进入 raw 检查的 ID 数: {checked_n}")

    log("")
    log(f"FOUND_IN_1_RAW_ONLY: {len(found)}")
    for item in found:
        rec = item["rec"]
        entry = item["entry"]
        log(
            f"  [FOUND_RAW] row={rec['row']} | ID={rec['sub_id']} "
            f"-> folder={entry.rel_path} | core_id={entry.core_id} | match={item['match']}"
        )

    log("")
    log(f"MISSING_IN_BOTH: {len(missing)}")
    for item in missing:
        rec = item["rec"]
        log(f"  [MISSING] row={rec['row']} | ID={rec['sub_id']} | raw_match={item['match']}")

    log("")
    log(f"AMBIGUOUS_IN_1_RAW: {len(ambiguous)}")
    for item in ambiguous:
        rec = item["rec"]
        log(
            f"  [AMBIGUOUS_RAW] row={rec['row']} | ID={rec['sub_id']} "
            f"| match={item['match']} | candidates={format_candidates(item['candidates'])}"
        )


def log_site_copy_plan(site_folder: str, final_rows: List[Dict[str, Any]]) -> None:
    log_subsection(f"[{site_folder}] COPY PLAN")

    processed_rows = [r for r in final_rows if r["status"] == "PROCESSED_FOUND"]
    raw_rows = [r for r in final_rows if r["status"] == "RAW_FOUND_ONLY"]
    no_action_rows = [r for r in final_rows if r["status"] not in ["PROCESSED_FOUND", "RAW_FOUND_ONLY"]]

    log(f"COPY_TO_6_RESULTS: {len(processed_rows)}")
    for r in processed_rows:
        log(f"  [TO_6_RESULTS] ID={r['sub_id']} | src={r['source_path']} | dst={r['dest_path']}")

    log("")
    log(f"COPY_TO_0_RAWDATA: {len(raw_rows)}")
    for r in raw_rows:
        log(f"  [TO_0_RAWDATA] ID={r['sub_id']} | src={r['source_path']} | dst={r['dest_path']}")

    log("")
    log(f"NO_ACTION: {len(no_action_rows)}")
    for r in no_action_rows:
        log(f"  [NO_ACTION] ID={r['sub_id']} | status={r['status']} | match={r['match_type']}")


# =========================
# 主流程
# =========================

def main():
    init_log()

    try:
        print("=" * 80)
        print("CBCP MRI QC scan/copy script")
        print("=" * 80)

        if EXECUTE_COPY:
            print("[MODE] EXECUTE_COPY=True：会真正复制文件夹。")
        else:
            print("[MODE] EXECUTE_COPY=False：DRY RUN，只扫描和写日志，不复制。")

        print(f"[LOG] 详细日志：{LOG_PATH}")
        print(f"[CSV] 扫描报告：{REPORT_CSV_PATH}")

        log_section("CBCP MRI QC scan/copy script")
        log(f"EXECUTE_COPY            : {EXECUTE_COPY}")
        log(f"MAX_SEARCH_DEPTH        : {MAX_SEARCH_DEPTH}")
        log(f"ALLOW_CONTAINS_FALLBACK : {ALLOW_CONTAINS_FALLBACK}")
        log(f"EXCEL_PATH              : {EXCEL_PATH}")
        log(f"PROJECT_DIR             : {PROJECT_DIR}")
        log(f"LOCAL_RAW_DIR           : {LOCAL_RAW_DIR}")
        log(f"LOCAL_RESULTS_DIR       : {LOCAL_RESULTS_DIR}")

        excel_path = Path(EXCEL_PATH)

        if not excel_path.exists():
            raise FileNotFoundError(f"指定的 Excel 不存在：{excel_path}")

        run_net_use()

        records = load_ids_from_excel(excel_path)

        if len(records) == 0:
            print("[ERROR] Excel 中没有读到任何 ID。")
            log("[ERROR] Excel 中没有读到任何 ID。")
            return

        records_by_site = defaultdict(list)
        for rec in records:
            records_by_site[rec["site_folder"]].append(rec)

        print(f"[EXCEL] 总 ID 数：{len(records)}")
        for site_folder, site_records in records_by_site.items():
            print(f"  {site_folder}: {len(site_records)}")

        all_results = []

        total_processed = 0
        total_raw = 0
        total_missing = 0
        total_ambiguous_processed = 0
        total_ambiguous_raw = 0
        total_copied = 0
        total_skipped_existing = 0
        total_copy_failed = 0

        log_section("[SITE-BY-SITE MATCH RESULT]")

        for site_folder in sorted(records_by_site.keys()):
            site_records = records_by_site[site_folder]

            #processed_parent = REMOTE_BASE / site_folder / "1_sMRI" / "3_sMRI_preprocessing"
            processed_parent = REMOTE_BASE / site_folder / "1_sMRI" / "2_sMRI_QCpass_T1_T2"
            #2_sMRI_QCpass_T1_T2
            raw_parent = REMOTE_BASE / site_folder / "1_sMRI" / "1_sMRI_data"

            processed_index = build_folder_index(processed_parent, MAX_SEARCH_DEPTH)
            raw_index = build_folder_index(raw_parent, MAX_SEARCH_DEPTH)

            log_section(f"SITE: {site_folder}")
            log_site_index_info(site_folder, processed_index, raw_index)

            # Step 1：所有 ID 先查 3_sMRI_preprocessing
            found_3 = []
            not_found_3 = []
            ambiguous_3 = []

            site_final_rows = []

            for rec in site_records:
                sub_id = rec["sub_id"]
                entry, match, candidates = find_id_folder(processed_index, sub_id)

                if entry is not None:
                    found_3.append({
                        "rec": rec,
                        "entry": entry,
                        "match": match,
                        "candidates": candidates,
                    })

                    final_row = {
                        "sheet": rec["sheet"],
                        "row": rec["row"],
                        "site_folder": site_folder,
                        "sub_id": sub_id,
                        "excel_raw_value": rec["cell_debug"].get("raw_value"),
                        "excel_python_type": rec["cell_debug"].get("python_type"),
                        "excel_number_format": rec["cell_debug"].get("number_format"),
                        "status": "PROCESSED_FOUND",
                        "match_type": match,
                        "matched_folder": entry.rel_path,
                        "matched_core_id": entry.core_id,
                        "source_path": str(entry.path),
                        "dest_path": str(LOCAL_RESULTS_DIR / site_folder / entry.name),
                        "action": "COPY_TO_6_RESULTS",
                        "copy_result": "DRY_RUN_NO_COPY",
                    }

                    site_final_rows.append(final_row)

                elif candidates:
                    ambiguous_3.append({
                        "rec": rec,
                        "match": match,
                        "candidates": candidates,
                    })

                    final_row = {
                        "sheet": rec["sheet"],
                        "row": rec["row"],
                        "site_folder": site_folder,
                        "sub_id": sub_id,
                        "excel_raw_value": rec["cell_debug"].get("raw_value"),
                        "excel_python_type": rec["cell_debug"].get("python_type"),
                        "excel_number_format": rec["cell_debug"].get("number_format"),
                        "status": "AMBIGUOUS_PROCESSED",
                        "match_type": match,
                        "matched_folder": "",
                        "matched_core_id": "",
                        "source_path": "",
                        "dest_path": "",
                        "action": "NO_ACTION_AMBIGUOUS",
                        "copy_result": "DRY_RUN_NO_COPY",
                    }

                    site_final_rows.append(final_row)

                else:
                    not_found_3.append({
                        "rec": rec,
                        "match": match,
                        "candidates": candidates,
                    })

            log_site_step1(
                site_folder=site_folder,
                total_n=len(site_records),
                found=found_3,
                not_found=not_found_3,
                ambiguous=ambiguous_3,
            )

            # Step 2：只有 3_sMRI_preprocessing 没找到的，才查 raw
            found_raw = []
            missing_both = []
            ambiguous_raw = []

            for item in not_found_3:
                rec = item["rec"]
                sub_id = rec["sub_id"]

                entry, match, candidates = find_id_folder(raw_index, sub_id)

                if entry is not None:
                    found_raw.append({
                        "rec": rec,
                        "entry": entry,
                        "match": match,
                        "candidates": candidates,
                    })

                    final_row = {
                        "sheet": rec["sheet"],
                        "row": rec["row"],
                        "site_folder": site_folder,
                        "sub_id": sub_id,
                        "excel_raw_value": rec["cell_debug"].get("raw_value"),
                        "excel_python_type": rec["cell_debug"].get("python_type"),
                        "excel_number_format": rec["cell_debug"].get("number_format"),
                        "status": "RAW_FOUND_ONLY",
                        "match_type": match,
                        "matched_folder": entry.rel_path,
                        "matched_core_id": entry.core_id,
                        "source_path": str(entry.path),
                        "dest_path": str(LOCAL_RAW_DIR / site_folder / entry.name),
                        "action": "COPY_TO_0_RAWDATA",
                        "copy_result": "DRY_RUN_NO_COPY",
                    }

                    site_final_rows.append(final_row)

                elif candidates:
                    ambiguous_raw.append({
                        "rec": rec,
                        "match": match,
                        "candidates": candidates,
                    })

                    final_row = {
                        "sheet": rec["sheet"],
                        "row": rec["row"],
                        "site_folder": site_folder,
                        "sub_id": sub_id,
                        "excel_raw_value": rec["cell_debug"].get("raw_value"),
                        "excel_python_type": rec["cell_debug"].get("python_type"),
                        "excel_number_format": rec["cell_debug"].get("number_format"),
                        "status": "AMBIGUOUS_RAW",
                        "match_type": match,
                        "matched_folder": "",
                        "matched_core_id": "",
                        "source_path": "",
                        "dest_path": "",
                        "action": "NO_ACTION_AMBIGUOUS",
                        "copy_result": "DRY_RUN_NO_COPY",
                    }

                    site_final_rows.append(final_row)

                else:
                    missing_both.append({
                        "rec": rec,
                        "match": match,
                        "candidates": candidates,
                    })

                    final_row = {
                        "sheet": rec["sheet"],
                        "row": rec["row"],
                        "site_folder": site_folder,
                        "sub_id": sub_id,
                        "excel_raw_value": rec["cell_debug"].get("raw_value"),
                        "excel_python_type": rec["cell_debug"].get("python_type"),
                        "excel_number_format": rec["cell_debug"].get("number_format"),
                        "status": "MISSING",
                        "match_type": f"processed_not_found; raw_{match}",
                        "matched_folder": "",
                        "matched_core_id": "",
                        "source_path": "",
                        "dest_path": "",
                        "action": "NO_ACTION",
                        "copy_result": "DRY_RUN_NO_COPY",
                    }

                    site_final_rows.append(final_row)

            log_site_step2(
                site_folder=site_folder,
                checked_n=len(not_found_3),
                found=found_raw,
                missing=missing_both,
                ambiguous=ambiguous_raw,
            )

            log_site_copy_plan(site_folder, site_final_rows)

            # 执行复制
            if EXECUTE_COPY:
                log_subsection(f"[{site_folder}] EXECUTE COPY")
                for r in site_final_rows:
                    if r["status"] not in ["PROCESSED_FOUND", "RAW_FOUND_ONLY"]:
                        continue

                    ok, msg = robocopy_folder(Path(r["source_path"]), Path(r["dest_path"]))
                    r["copy_result"] = msg

                    if ok:
                        if msg == "SKIPPED_DEST_EXISTS":
                            total_skipped_existing += 1
                        else:
                            total_copied += 1
                    else:
                        total_copy_failed += 1

            all_results.extend(site_final_rows)

            site_processed = len(found_3)
            site_raw = len(found_raw)
            site_missing = len(missing_both)
            site_ambiguous_processed = len(ambiguous_3)
            site_ambiguous_raw = len(ambiguous_raw)

            total_processed += site_processed
            total_raw += site_raw
            total_missing += site_missing
            total_ambiguous_processed += site_ambiguous_processed
            total_ambiguous_raw += site_ambiguous_raw

            print(
                f"[{site_folder}] Excel={len(site_records)} | "
                f"3_preprocessing found={site_processed} | "
                f"raw found={site_raw} | "
                f"missing={site_missing} | "
                f"ambiguous_3={site_ambiguous_processed} | "
                f"ambiguous_raw={site_ambiguous_raw}"
            )

        # 写 CSV
        with open(REPORT_CSV_PATH, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "sheet",
                    "row",
                    "site_folder",
                    "sub_id",
                    "excel_raw_value",
                    "excel_python_type",
                    "excel_number_format",
                    "status",
                    "match_type",
                    "matched_folder",
                    "matched_core_id",
                    "source_path",
                    "dest_path",
                    "action",
                    "copy_result",
                ]
            )
            writer.writeheader()
            writer.writerows(all_results)

        log_section("[TOTAL SUMMARY]")
        log(f"Excel 总 ID 数              : {len(records)}")
        log(f"已预处理，去 6_results      : {total_processed}")
        log(f"只有原始数据，去 0_rawdata  : {total_raw}")
        log(f"processed 多候选，未复制    : {total_ambiguous_processed}")
        log(f"raw 多候选，未复制          : {total_ambiguous_raw}")
        log(f"NAS 两边都没找到            : {total_missing}")

        if EXECUTE_COPY:
            log(f"实际复制成功                : {total_copied}")
            log(f"目标已存在，跳过            : {total_skipped_existing}")
            log(f"复制失败                    : {total_copy_failed}")
        else:
            log("当前是 DRY RUN，没有真正复制。确认无误后，把 EXECUTE_COPY 改成 True。")

        print()
        print("=" * 80)
        print("[TOTAL SUMMARY]")
        print(f"Excel 总 ID 数              : {len(records)}")
        print(f"已预处理，去 6_results      : {total_processed}")
        print(f"只有原始数据，去 0_rawdata  : {total_raw}")
        print(f"processed 多候选，未复制    : {total_ambiguous_processed}")
        print(f"raw 多候选，未复制          : {total_ambiguous_raw}")
        print(f"NAS 两边都没找到            : {total_missing}")

        if EXECUTE_COPY:
            print(f"实际复制成功                : {total_copied}")
            print(f"目标已存在，跳过            : {total_skipped_existing}")
            print(f"复制失败                    : {total_copy_failed}")
        else:
            print("当前是 DRY RUN，没有真正复制。")

        print(f"详细日志：{LOG_PATH}")
        print(f"CSV 报告：{REPORT_CSV_PATH}")
        print("=" * 80)

    finally:
        close_log()


if __name__ == "__main__":
    main()