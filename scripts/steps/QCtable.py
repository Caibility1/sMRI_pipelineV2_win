from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from datetime import datetime, date
import re

file_path = r"D:\University\master\QC2026\2026\0422_CBCP\CBCP_QC.xlsx"

src_sheet_name = "RealTimeQC"
tgt_sheet_name = "skd_Allqc"

target_row_start = 270
target_row_end = 306


def norm_header(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = s.replace(" ", "").replace("_", "")
    s = s.replace("？", "?")
    return s


def norm_subnum(v):
    if v is None:
        return ""
    s = "".join(ch for ch in str(v).strip() if ch.isdigit())
    if s == "":
        return ""
    s = s.lstrip("0")
    return s if s else "0"


def find_header_row(ws, required_headers, search_rows=10):
    required = [norm_header(x) for x in required_headers]
    for r in range(1, search_rows + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        row_norm = [norm_header(x) for x in row_vals]
        if all(x in row_norm for x in required):
            return r
    return None


def build_col_map(ws, header_row):
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip() != "":
            col_map[norm_header(v)] = c
    return col_map


def parse_date(v):
    if v is None or v == "":
        return None

    if isinstance(v, datetime):
        return v.date()

    if isinstance(v, date):
        return v

    if isinstance(v, (int, float)):
        try:
            d = from_excel(v)
            if isinstance(d, datetime):
                return d.date()
            if isinstance(d, date):
                return d
        except:
            pass

    s = str(v).strip()
    if s == "":
        return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt).date()
        except:
            pass

    m = re.match(r"^\s*(\d{4})\D+(\d{1,2})\D+(\d{1,2})\s*$", s)
    if m:
        y, mo, d = map(int, m.groups())
        return date(y, mo, d)

    return None


def calc_month_age(birth_v, scan_v):
    birth = parse_date(birth_v)
    scan = parse_date(scan_v)
    if birth is None or scan is None:
        return None

    months = (scan.year - birth.year) * 12 + (scan.month - birth.month)
    if scan.day < birth.day:
        months -= 1
    return months


def norm_gender(v):
    if v is None:
        return None
    s = str(v).strip().lower()

    if s in {"男", "m", "male"}:
        return "M"
    if s in {"女", "f", "female"}:
        return "F"

    return str(v).strip()


def norm_status(v):
    if v is None:
        return None

    s = str(v).strip()
    if s == "":
        return None

    s2 = s.lower().replace(" ", "").replace("，", ",").replace("？", "").replace("?", "")

    if s2 in {"睡眠", "sleep", "asleep", "1", "是", "y", "yes"}:
        return "睡眠"
    if s2 in {"清醒", "awake", "0", "否", "n", "no"}:
        return "清醒"

    if "睡" in s2:
        return "睡眠"
    if "醒" in s2:
        return "清醒"

    return s.strip()


wb = load_workbook(file_path)
ws_src = wb[src_sheet_name]
ws_tgt = wb[tgt_sheet_name]

src_required = ["subnum", "gender", "scan_date", "birth_date", "月龄", "是否睡眠？"]
tgt_required = ["subnum", "gender", "birthdate", "scandate", "age", "status"]

src_header_row = find_header_row(ws_src, src_required)
tgt_header_row = find_header_row(ws_tgt, tgt_required)

if src_header_row is None:
    raise ValueError("RealTimeQC 没找到表头。")
if tgt_header_row is None:
    raise ValueError("skd_Allqc 没找到表头。")

src_col = build_col_map(ws_src, src_header_row)
tgt_col = build_col_map(ws_tgt, tgt_header_row)

src_subnum_col = src_col[norm_header("subnum")]
src_gender_col = src_col[norm_header("gender")]
src_scan_date_col = src_col[norm_header("scan_date")]
src_birth_date_col = src_col[norm_header("birth_date")]
src_age_col = src_col[norm_header("月龄")]
src_status_col = src_col[norm_header("是否睡眠？")]

tgt_subnum_col = tgt_col[norm_header("subnum")]
tgt_gender_col = tgt_col[norm_header("gender")]
tgt_birthdate_col = tgt_col[norm_header("birthdate")]
tgt_scandate_col = tgt_col[norm_header("scandate")]
tgt_age_col = tgt_col[norm_header("age")]
tgt_status_col = tgt_col[norm_header("status")]

id_to_src_row = {}
duplicate_ids = []

for r in range(src_header_row + 1, ws_src.max_row + 1):
    raw_id = ws_src.cell(r, src_subnum_col).value
    sid = norm_subnum(raw_id)
    if sid == "":
        continue

    if sid in id_to_src_row:
        duplicate_ids.append(sid)
    else:
        id_to_src_row[sid] = r

if duplicate_ids:
    print("以下标准化后的 subnum 在 RealTimeQC 中重复，默认使用第一次出现：")
    for x in sorted(set(duplicate_ids)):
        print("  ", x)

filled_count = 0
not_found = []

for r in range(target_row_start, target_row_end + 1):
    raw_id = ws_tgt.cell(r, tgt_subnum_col).value
    sid = norm_subnum(raw_id)

    if sid == "":
        print(f"第 {r} 行 subnum 为空，跳过")
        continue

    if sid not in id_to_src_row:
        not_found.append((r, raw_id))
        print(f"没找到 subnum: 第 {r} 行 -> {raw_id}")
        continue

    src_r = id_to_src_row[sid]

    src_gender = norm_gender(ws_src.cell(src_r, src_gender_col).value)
    src_birth = parse_date(ws_src.cell(src_r, src_birth_date_col).value)
    src_scan = parse_date(ws_src.cell(src_r, src_scan_date_col).value)
    src_age_raw = ws_src.cell(src_r, src_age_col).value
    src_status = norm_status(ws_src.cell(src_r, src_status_col).value)

    if src_age_raw is None or str(src_age_raw).strip() == "":
        src_age = calc_month_age(src_birth, src_scan)
    else:
        try:
            src_age = int(float(src_age_raw))
        except:
            src_age = calc_month_age(src_birth, src_scan)

    ws_tgt.cell(r, tgt_gender_col).value = src_gender
    ws_tgt.cell(r, tgt_birthdate_col).value = src_birth
    ws_tgt.cell(r, tgt_scandate_col).value = src_scan
    ws_tgt.cell(r, tgt_age_col).value = src_age
    ws_tgt.cell(r, tgt_status_col).value = src_status

    ws_tgt.cell(r, tgt_birthdate_col).number_format = "yyyy-mm-dd"
    ws_tgt.cell(r, tgt_scandate_col).number_format = "yyyy-mm-dd"

    print(
        f"已填充 第 {r} 行, subnum={raw_id} -> "
        f"{src_gender}, {src_birth}, {src_scan}, {src_age}, {src_status}"
    )
    filled_count += 1

wb.save(file_path)

print("\n===== 完成 =====")
print(f"成功填充: {filled_count} 行")
print(f"已直接覆盖原文件: {file_path}")

if not_found:
    print("\n以下仍未找到：")
    for r, raw_id in not_found:
        print(f"  第 {r} 行: {raw_id}")